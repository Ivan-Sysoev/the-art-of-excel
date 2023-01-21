from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from progress.bar import IncrementalBar
from os.path import abspath

image = "IMG_0003 (1).jpg"

im = Image.open(image)
rgb_code = im.convert("RGB")
width, height = im.size
bar = IncrementalBar("Countdown", max = width * height)

wb = openpyxl.Workbook()
ws = wb.active

for n in range(width):
    for i in range(height):
        color = "%02x%02x%02x" % rgb_code.getpixel((i,n))
        colorFill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        cell = ws.cell(row = n + 1, column = i + 1)
        cell.fill = colorFill
        ws.column_dimensions[cell.column_letter].width = 2.9
        bar.next()

filename = image.rsplit(".", 1)[0] + ".xlsx"
wb.save(filename=filename)
print(f"\nOPEN the result: {abspath(filename)}")